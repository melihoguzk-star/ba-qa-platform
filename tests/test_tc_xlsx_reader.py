"""
tests/test_tc_xlsx_reader.py

TC XLSX Reader modülü için test suite.
pipeline/tc_xlsx_reader.py → read_tc_xlsx(), _map_headers(), _find_header_row(),
                              _extract_cover_meta(), _is_valid_row()
"""
import pytest
from io import BytesIO


# ── Yardımcı: minimal workbook factory ────────────────────────────────────────

def _make_minimal_xlsx(headers: list[str], rows: list[list], title: str = "TestSheet") -> bytes:
    """Tek content sheet'li minimal XLSX oluşturur."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Import ────────────────────────────────────────────────────────────────────

def test_import():
    """Modül import edilebilmeli."""
    from pipeline.tc_xlsx_reader import (
        read_tc_xlsx, COLUMN_ALIASES, SKIP_SHEETS,
        REQUIRED_FOR_VALID_ROW, _ALIAS_LOOKUP,
    )
    assert "testcase_name" in COLUMN_ALIASES
    assert "Cover" in SKIP_SHEETS


# ── COLUMN_ALIASES ve _ALIAS_LOOKUP ──────────────────────────────────────────

class TestAliasLookup:
    """_ALIAS_LOOKUP reverse mapping testleri."""

    def test_testcase_alias(self):
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["TESTCASE"] == "testcase_name"

    def test_test_case_with_space_alias(self):
        """'TEST CASE' (boşluklu) da testcase_name'e map edilmeli."""
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["TEST CASE"] == "testcase_name"

    def test_priority_typo(self):
        """PRIORTY (typo) → priority."""
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["PRIORTY"] == "priority"
        assert _ALIAS_LOOKUP["PRIORITY"] == "priority"

    def test_existance_typo(self):
        """EXISTANCE (typo) → existence."""
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["EXISTANCE"] == "existence"

    def test_existance_variants(self):
        """EXISTANCE 2/3/4 → existence."""
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["EXISTANCE 2"] == "existence"
        assert _ALIAS_LOOKUP["EXISTANCE 3"] == "existence"
        assert _ALIAS_LOOKUP["EXISTANCE 4"] == "existence"

    def test_lowercase_c_alias(self):
        """'c' (lowercase) → existence (Çekirdek ile ödeme sheet typo)."""
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["C"] == "existence"

    def test_created_by_variants(self):
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["CREATED BY"] == "created_by"
        assert _ALIAS_LOOKUP["CREATED By".upper()] == "created_by"

    def test_test_case_id(self):
        from pipeline.tc_xlsx_reader import _ALIAS_LOOKUP
        assert _ALIAS_LOOKUP["TEST CASE ID"] == "test_case_id"


# ── Skip sheets ────────────────────────────────────────────────────────────────

class TestSkipSheets:
    """Skip edilecek sheet'lerin tespiti."""

    def test_cover_skipped(self, sample_tc_xlsx):
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        result = read_tc_xlsx(sample_tc_xlsx)
        sheet_names = [s["sheet_name"] for s in result["sheets"]]
        assert "Cover" not in sheet_names
        assert "Cover" in result["skip_sheets"]

    def test_data_skipped(self, sample_tc_xlsx):
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        result = read_tc_xlsx(sample_tc_xlsx)
        sheet_names = [s["sheet_name"] for s in result["sheets"]]
        assert "DATA" not in sheet_names
        assert "DATA" in result["skip_sheets"]

    def test_revision_changes_skipped(self, sample_tc_xlsx):
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        result = read_tc_xlsx(sample_tc_xlsx)
        sheet_names = [s["sheet_name"] for s in result["sheets"]]
        assert "Revision Changes" not in sheet_names
        assert "Revision Changes" in result["skip_sheets"]

    def test_content_sheets_present(self, sample_tc_xlsx):
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        result = read_tc_xlsx(sample_tc_xlsx)
        sheet_names = [s["sheet_name"] for s in result["sheets"]]
        assert "Login" in sheet_names
        assert "Sepet" in sheet_names


# ── Cover meta ────────────────────────────────────────────────────────────────

class TestCoverMeta:
    """Cover sheet'ten meta bilgi çıkarma."""

    def test_document_code(self, sample_tc_raw):
        assert sample_tc_raw["meta"]["document_code"] == "LDS-PDQA-MT-KD-1"

    def test_project_name(self, sample_tc_raw):
        assert sample_tc_raw["meta"]["project_name"] == "Kahve Dünyası"

    def test_created_by(self, sample_tc_raw):
        assert sample_tc_raw["meta"]["created_by"] == "Salih Diken"

    def test_create_date(self, sample_tc_raw):
        assert sample_tc_raw["meta"]["create_date"] == "24.04.2024"

    def test_no_cover_returns_empty_meta(self):
        """Cover sheet olmadan meta boş dict dönmeli."""
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active; ws.title = "Content"
        ws.cell(row=1, column=1, value="TESTCASE")
        ws.cell(row=2, column=1, value="Bir test")
        buf = BytesIO(); wb.save(buf)
        result = read_tc_xlsx(buf.getvalue())
        assert result["meta"] == {}


# ── Header mapping ─────────────────────────────────────────────────────────────

class TestHeaderMapping:
    """Header eşleme ve varyant desteği."""

    def test_variant1_testcase(self, sample_tc_raw):
        """TESTCASE kolonu testcase_name'e map edilmeli."""
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("testcase_name") == "Login TC Adı"

    def test_variant2_test_case_with_space(self, sample_tc_raw):
        """TEST CASE (boşluklu) kolonu da testcase_name'e map edilmeli."""
        sepet = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Sepet")
        assert sepet["rows"][0].get("testcase_name") == "Sepet Testi"

    def test_priority_typo_mapped(self, sample_tc_raw):
        """PRIORTY (typo) → priority."""
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("priority") == "HIGH"

    def test_existance_mapped(self, sample_tc_raw):
        """EXISTANCE → existence."""
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("existence") == "New"

    def test_existance_variants_mapped(self, sample_tc_raw):
        """EXISTANCE 2/3/4 → existence (pinarÜrünDetay sheet)."""
        pinar = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "pinarÜrünDetay")
        # EXISTANCE 2/3/4 all map to 'existence' — last write wins in row dict
        row = pinar["rows"][0]
        assert "existence" in row

    def test_otomasyon_mapped(self, sample_tc_raw):
        """OTOMASYON kolonu doğru eşlenmeli."""
        reg = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Regresyon Seti")
        assert reg["rows"][0].get("otomasyon") == "Otomasyon Hazır"

    def test_unknown_column_prefixed(self):
        """Bilinmeyen kolon 'unknown_' prefix ile saklanmalı."""
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        xlsx = _make_minimal_xlsx(
            ["TESTCASE", "TEST STEPS", "EXPECTED RESULT", "BILINMEYEN_KOLON"],
            [["Test adı", "Adım 1", "Beklenen", "Değer"]]
        )
        result = read_tc_xlsx(xlsx)
        row = result["sheets"][0]["rows"][0]
        assert any(k.startswith("unknown_") for k in row), f"Unknown col not found: {row}"


# ── Boş satır filtreleme ───────────────────────────────────────────────────────

class TestEmptyRowFiltering:
    """Boş satırların doğru filtrelenmesi."""

    def test_login_has_2_valid_rows(self, sample_tc_raw):
        """Login sheet'te 2 geçerli satır olmalı (3 boş satır filtre edilmeli)."""
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert len(login["rows"]) == 2

    def test_empty_only_sheet_skipped(self):
        """Sadece header + boş satır olan sheet atlanmalı veya 0 row döndürmeli."""
        from pipeline.tc_xlsx_reader import read_tc_xlsx
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active; ws.title = "EmptySheet"
        for i, h in enumerate(["TESTCASE", "TEST STEPS", "EXPECTED RESULT"], 1):
            ws.cell(row=1, column=i, value=h)
        # Satır 2-100: tamamen boş
        buf = BytesIO(); wb.save(buf)
        result = read_tc_xlsx(buf.getvalue())
        sheet = next((s for s in result["sheets"] if s["sheet_name"] == "EmptySheet"), None)
        if sheet:
            assert len(sheet["rows"]) == 0

    def test_is_valid_row_true(self):
        """En az bir REQUIRED alan dolu → geçerli."""
        from pipeline.tc_xlsx_reader import _is_valid_row
        assert _is_valid_row({"testcase_name": "Test", "test_steps": None}) is True

    def test_is_valid_row_false_all_empty(self):
        """Tüm alanlar None → geçersiz."""
        from pipeline.tc_xlsx_reader import _is_valid_row
        assert _is_valid_row({"testcase_name": None, "test_steps": None, "expected_result": None}) is False

    def test_is_valid_row_false_blank_strings(self):
        """Tüm alanlar boş string → geçersiz."""
        from pipeline.tc_xlsx_reader import _is_valid_row
        assert _is_valid_row({"testcase_name": "  ", "test_steps": "", "expected_result": "   "}) is False


# ── Multiline test steps ───────────────────────────────────────────────────────

class TestMultilineContent:
    """Multiline hücre içeriği korunmalı."""

    def test_newline_preserved_in_steps(self, sample_tc_raw):
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        steps = login["rows"][0].get("test_steps", "")
        assert "\n" in steps, f"Newline not found in: {repr(steps)}"

    def test_step_content(self, sample_tc_raw):
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        steps = login["rows"][0].get("test_steps", "")
        assert "Uygulama açılır" in steps


# ── _find_header_row ───────────────────────────────────────────────────────────

class TestFindHeaderRow:
    """Header satırı tespiti."""

    def test_header_at_row_1(self):
        from pipeline.tc_xlsx_reader import _find_header_row
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1, value="TESTCASE")
        ws.cell(row=1, column=2, value="TEST STEPS")
        ws.cell(row=1, column=3, value="EXPECTED RESULT")
        ws.cell(row=2, column=1, value="Test adı")
        assert _find_header_row(ws) == 1

    def test_header_at_row_3(self):
        """İlk 2 satır boşsa satır 3'te header bulunmalı."""
        from pipeline.tc_xlsx_reader import _find_header_row
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        # Satır 1-2: boş
        ws.cell(row=3, column=1, value="TESTCASE")
        ws.cell(row=3, column=2, value="TEST STEPS")
        ws.cell(row=3, column=3, value="EXPECTED RESULT")
        assert _find_header_row(ws) == 3

    def test_no_header_returns_none(self):
        """Hiç header yoksa None döndürmeli."""
        from pipeline.tc_xlsx_reader import _find_header_row
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        # Tamamen boş sheet
        assert _find_header_row(ws) is None


# ── _cell_to_str ───────────────────────────────────────────────────────────────

class TestCellToStr:
    """Hücre değeri dönüşüm helper."""

    def test_none_returns_none(self):
        from pipeline.tc_xlsx_reader import _cell_to_str
        assert _cell_to_str(None) is None

    def test_empty_string_returns_none(self):
        from pipeline.tc_xlsx_reader import _cell_to_str
        assert _cell_to_str("   ") is None

    def test_string_stripped(self):
        from pipeline.tc_xlsx_reader import _cell_to_str
        assert _cell_to_str("  test  ") == "test"

    def test_datetime_formatted(self):
        import datetime
        from pipeline.tc_xlsx_reader import _cell_to_str
        d = datetime.date(2024, 5, 20)
        assert _cell_to_str(d) == "20.05.2024"

    def test_integer_to_string(self):
        from pipeline.tc_xlsx_reader import _cell_to_str
        assert _cell_to_str(42) == "42"

    def test_newline_preserved(self):
        from pipeline.tc_xlsx_reader import _cell_to_str
        assert _cell_to_str("line1\nline2") == "line1\nline2"


# ── End-to-end: tam fixture üzerinde ─────────────────────────────────────────

class TestEndToEnd:
    """Gerçekçi fixture ile uçtan uca reader testi."""

    def test_sheet_count(self, sample_tc_raw):
        """Skip sheet'ler hariç 4 content sheet olmalı."""
        assert len(sample_tc_raw["sheets"]) == 4

    def test_skip_count(self, sample_tc_raw):
        """Cover + DATA + Revision Changes = 3 skip."""
        assert len(sample_tc_raw["skip_sheets"]) == 3

    def test_all_content_sheets_have_rows(self, sample_tc_raw):
        """Her content sheet'te en az 1 geçerli satır olmalı."""
        for sheet in sample_tc_raw["sheets"]:
            assert len(sheet["rows"]) >= 1, f"Sheet '{sheet['sheet_name']}' has no rows"

    def test_expected_result_present(self, sample_tc_raw):
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("expected_result") == "Başarılı giriş beklenir"

    def test_channel_present(self, sample_tc_raw):
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("channel") == "MOBILE"

    def test_test_area_present(self, sample_tc_raw):
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("test_area") == "Login Ekranı"

    def test_test_case_id_present(self, sample_tc_raw):
        login = next(s for s in sample_tc_raw["sheets"] if s["sheet_name"] == "Login")
        assert login["rows"][0].get("test_case_id") == "TC_LDS_KD_LOGIN_0001"
