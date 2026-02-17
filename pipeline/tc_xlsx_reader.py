"""
Loodos Test Case XLSX okuyucu.
openpyxl kullanarak sheet'leri okur, header'ları eşler, boş satırları filtreler.
"""
from __future__ import annotations

from io import BytesIO
from typing import Optional

import openpyxl

# ---------------------------------------------------------------------------
# Column alias mapping (canonical → list of possible header strings)
# Case-insensitive matching; typos are intentional (PRIORTY, EXISTANCE)
# ---------------------------------------------------------------------------
COLUMN_ALIASES: dict[str, list[str]] = {
    "test_case_id":       ["TEST CASE ID"],
    "testcase_name":      ["TESTCASE", "TEST CASE"],   # two variants!
    "test_steps":         ["TEST STEPS"],
    "expected_result":    ["EXPECTED RESULT"],
    "priority":           ["PRIORTY", "PRIORITY"],     # typo variant
    "channel":            ["CHANNEL"],
    "testcase_type":      ["TESTCASE TYPE"],
    "user_type":          ["USER TYPE"],
    "test_area":          ["TEST AREA"],
    "test_scenario":      ["TEST SCENARIO"],
    "existence":          ["EXISTANCE", "EXISTANCE 2", "EXISTANCE 3", "EXISTANCE 4", "C"],
    "date":               ["DATE"],
    "app_bundle":         ["APP BUNDLE"],
    "br_id":              ["BR ID"],
    "tr_id":              ["TR ID"],
    "precondition":       ["PRECONDITION"],
    "test_data":          ["TEST DATA"],
    "postcondition":      ["POSTCONDITION"],
    "actual_result":      ["ACTUAL RESULT"],
    "status":             ["STATUS"],
    "regression":         ["REGRESSION CASE"],
    "main_business_case": ["MAIN BUSINESS CASE"],
    "comments":           ["COMMENTS"],
    "created_by":         ["CREATED BY", "CREATED By"],
    "module":             ["MODULE"],
    "otomasyon":          ["OTOMASYON"],
}

# Reverse lookup: UPPER_HEADER → canonical_name
_ALIAS_LOOKUP: dict[str, str] = {
    alias.upper(): canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}

# Sheets to skip (meta/reference sheets, not TC content)
SKIP_SHEETS: set[str] = {"Cover", "Revision Changes", "DATA"}

# At least one of these canonical fields must be non-empty for a row to be valid
REQUIRED_FOR_VALID_ROW: list[str] = ["testcase_name", "test_steps", "expected_result"]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def read_tc_xlsx(file_content: bytes) -> dict:
    """
    Test Case XLSX dosyasını okur.

    Args:
        file_content: XLSX dosyasının byte içeriği

    Returns:
        {
            "meta": {...},          # Cover sheet'ten çıkarılan bilgiler
            "sheets": [...],        # Her content sheet'in raw datası
            "skip_sheets": [...]    # Atlanan sheet isimleri
        }
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    meta: dict = {}
    sheets: list[dict] = []
    skipped: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip meta sheets
        if sheet_name in SKIP_SHEETS:
            skipped.append(sheet_name)
            if sheet_name == "Cover":
                meta = _extract_cover_meta(ws)
            continue

        header_row = _find_header_row(ws)
        if header_row is None:
            skipped.append(sheet_name)
            continue

        col_map = _map_headers(ws, header_row)
        rows = _read_data_rows(ws, header_row, col_map)

        sheets.append({
            "sheet_name": sheet_name,
            "header_row": header_row,
            "col_map": col_map,
            "rows": rows,
        })

    return {
        "meta": meta,
        "sheets": sheets,
        "skip_sheets": skipped,
    }


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _extract_cover_meta(ws) -> dict:
    """
    Cover sheet'ten proje bilgilerini çıkarır.

    Bilinen layout:
      C2: "Document Code"  → D2: değer
      C3: "Project Name"   → D3: değer
      C4: "Created By"     → D4: değer
      C5: "Create Date"    → D5: değer
    """
    meta: dict = {}

    label_to_key = {
        "document code": "document_code",
        "project name":  "project_name",
        "created by":    "created_by",
        "create date":   "create_date",
    }

    # Search C2:D20 range for known labels
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=8, values_only=True):
        for i, cell_val in enumerate(row):
            if cell_val is None:
                continue
            label = str(cell_val).strip().lower()
            if label in label_to_key and i + 1 < len(row):
                value = row[i + 1]
                if value is not None:
                    meta[label_to_key[label]] = _cell_to_str(value)

    return meta


def _find_header_row(ws, max_search: int = 10) -> Optional[int]:
    """
    İlk max_search satır içinde, COLUMN_ALIASES'taki herhangi bir header değerini
    içeren ilk satırı döndürür.

    Fallback: en az 3 dolu hücre olan ilk satır.
    Returns: 1-based row number, or None if not found.
    """
    all_aliases_upper = set(_ALIAS_LOOKUP.keys())

    for row_idx in range(1, max_search + 1):
        row_values = [
            ws.cell(row=row_idx, column=c).value
            for c in range(1, ws.max_column + 1)
        ]
        non_empty = [v for v in row_values if v is not None and str(v).strip()]

        if not non_empty:
            continue

        # Check if any cell matches a known alias
        for val in non_empty:
            if str(val).strip().upper() in all_aliases_upper:
                return row_idx

        # Fallback: row with >= 3 non-empty cells might be a header
        if len(non_empty) >= 3:
            return row_idx

    return None


def _map_headers(ws, header_row: int) -> dict[int, str]:
    """
    Header satırındaki her hücreyi COLUMN_ALIASES kullanarak canonical isme eşler.

    Returns:
        {col_index (1-based): canonical_name}
        Eşleşmeyen kolonlar: {col_index: "unknown_ORIGINAL_NAME"}
    """
    col_map: dict[int, str] = {}

    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col_idx).value
        if cell_val is None:
            continue

        header_upper = str(cell_val).strip().upper()
        if not header_upper:
            continue

        canonical = _ALIAS_LOOKUP.get(header_upper)
        if canonical:
            col_map[col_idx] = canonical
        else:
            col_map[col_idx] = f"unknown_{str(cell_val).strip()}"

    return col_map


def _read_data_rows(ws, header_row: int, col_map: dict[int, str]) -> list[dict]:
    """
    Header satırının altındaki veri satırlarını okur.
    Boş satırları filtreler (REQUIRED_FOR_VALID_ROW kriterine göre).
    """
    rows: list[dict] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_dict: dict = {}

        for col_idx, canonical in col_map.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            row_dict[canonical] = _cell_to_str(cell.value)

        if _is_valid_row(row_dict):
            rows.append(row_dict)

    return rows


def _is_valid_row(row_dict: dict) -> bool:
    """
    Satırın geçerli bir TC olup olmadığını kontrol eder.
    REQUIRED_FOR_VALID_ROW'daki alanlardan en az biri dolu olmalı.
    """
    for field in REQUIRED_FOR_VALID_ROW:
        val = row_dict.get(field)
        if val is not None and str(val).strip():
            return True
    return False


def _cell_to_str(value) -> Optional[str]:
    """
    Hücre değerini stringe çevirir.
    - None → None
    - datetime → "DD.MM.YYYY"
    - float/int → string (Excel sayısal tarihlerini önle)
    - str → strip (newline'ları koru)
    """
    if value is None:
        return None

    import datetime
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%d.%m.%Y")

    text = str(value).strip()
    return text if text else None
