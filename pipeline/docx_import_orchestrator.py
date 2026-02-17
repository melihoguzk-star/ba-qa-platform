"""
DOCX Import Orchestrator — Adım 3

Upload edilen DOCX dosyasının şablon tipini otomatik algılar,
uygun parser'ı seçer ve confidence score hesaplar.

Desteklenen şablon tipleri:
  loodos_ba_bullet  — Kahve Dünyası formatı (nested bullet list, H2=ekran, H3=alt bölüm)
  loodos_ba_table   — Loodos BA tablo formatı (Tarih|Yazar|... tablosu, 1.1 Amaç, vb.)
  loodos_test_case  — Loodos TC Excel formatı (multi-sheet, TESTCASE/TEST STEPS headers)
  generic           — Genel doküman (rule-based heading parser fallback)
  generic_xlsx      — Bilinmeyen XLSX formatı
"""
import logging
from io import BytesIO
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

# Loodos BA bullet formatı için H3 işaret başlıkları
_LOODOS_BULLET_MARKERS = {"Açıklama", "İş Akışı", "Tasarım Dosyaları"}


def _tr_lower(s: str) -> str:
    """Turkish-aware lowercase: İ→i, I→ı, diğerleri standart lower()."""
    return s.replace('İ', 'i').replace('I', 'ı').lower()


class DocxImportOrchestrator:
    """
    DOCX dosyasını okur, şablon tipini algılar, uygun parser ile parse eder.

    Kullanım:
        orchestrator = DocxImportOrchestrator()
        result = orchestrator.import_docx(file_bytes)
    """

    def import_docx(
        self,
        file_content: bytes,
        doc_type: str = "auto",
        use_ai_fallback: bool = True,
        gemini_key: str = "",
    ) -> Dict[str, Any]:
        """
        DOCX bytes'ını parse edip sonuç dict'i döndür.

        Args:
            file_content:    .docx dosyasının binary içeriği
            doc_type:        "auto" | "ba" | "qa" — "auto" ise şablona göre seçilir
            use_ai_fallback: Düşük confidence'ta AI fallback kullanılsın mı?
            gemini_key:      AI fallback için Gemini API key (opsiyonel)

        Returns:
            {
                "success":      bool,
                "doc_type":     "ba" | "qa",
                "template":     "loodos_ba_bullet" | "loodos_ba_table" | "generic",
                "content_json": {...},
                "confidence":   0.0–1.0,
                "stats":        {"headings": int, "screens": int, "list_items": int,
                                 "paragraphs": int, "tables": int, "links": int},
                "warnings":     [str, ...]
            }
        """
        from pipeline.document_reader import read_docx_structured

        elements = read_docx_structured(file_content)
        stats    = self._calculate_stats(elements)
        template = self._detect_template(elements, stats)

        effective_doc_type = doc_type if doc_type != "auto" else "ba"

        if template in ("loodos_ba_bullet", "loodos_ba_table"):
            from pipeline.ba_docx_parser import BADocxParser
            content_json = BADocxParser().parse(elements)
            confidence   = self._calculate_confidence(content_json, stats)
        else:
            # Generic fallback — heading-based text parser
            try:
                from pipeline.document_reader import read_docx
                from pipeline.document_parser_v2 import parse_text_to_json
                text         = read_docx(file_content)
                content_json = parse_text_to_json(text, effective_doc_type)
                confidence   = 0.4
            except Exception as exc:
                logger.warning("Generic parser failed: %s", exc)
                content_json = {"ekranlar": [], "meta": {}, "linkler": {}}
                confidence   = 0.1

        # AI fallback — düşük confidence için (ileride implement edilecek)
        if confidence < 0.5 and use_ai_fallback and gemini_key:
            logger.info("AI fallback triggered (confidence=%.2f)", confidence)

        warnings = self._generate_warnings(content_json, stats)

        return {
            "success":      confidence > 0.3,
            "doc_type":     effective_doc_type,
            "template":     template,
            "content_json": content_json,
            "confidence":   confidence,
            "stats":        stats,
            "warnings":     warnings,
        }

    # ------------------------------------------------------------------
    # Şablon tespiti
    # ------------------------------------------------------------------

    def _detect_template(self, elements: List[Dict], stats: Dict) -> str:
        """
        Element listesinden şablon tipini algıla.

        Kural 1 — loodos_ba_bullet:
            H3 başlıkları arasında ≥2 Loodos işaret başlığı varsa
            (Açıklama, İş Akışı, Tasarım Dosyaları) VE list_items > tables

        Kural 2 — loodos_ba_table:
            Tablolar > 3 (tablo ağırlıklı doküman)

        Kural 3 — generic:
            Hiçbirine uymadı
        """
        h3_texts = {
            e["text"]
            for e in elements
            if e["type"] == "heading" and e.get("level") == 3
        }

        # Turkish-aware case-insensitive karşılaştırma
        h3_lower      = {_tr_lower(t.strip()) for t in h3_texts}
        markers_lower = {_tr_lower(m) for m in _LOODOS_BULLET_MARKERS}
        matched_count = sum(1 for m in markers_lower if m in h3_lower)

        if matched_count >= 2 and stats["list_items"] > stats.get("tables", 0):
            return "loodos_ba_bullet"

        if stats.get("tables", 0) > 3:
            return "loodos_ba_table"

        return "generic"

    # ------------------------------------------------------------------
    # İstatistik hesabı
    # ------------------------------------------------------------------

    def _calculate_stats(self, elements: List[Dict]) -> Dict:
        """Element listesinden doküman istatistiklerini hesapla."""
        headings = [e for e in elements if e["type"] == "heading"]
        return {
            "headings":   len(headings),
            "screens":    len([e for e in headings if e.get("level") == 2]),
            "list_items": len([e for e in elements if e["type"] == "list_item"]),
            "paragraphs": len([e for e in elements if e["type"] == "paragraph"]),
            "tables":     len([e for e in elements if e["type"] == "table"]),
            "links":      sum(len(e.get("links", [])) for e in elements),
        }

    # ------------------------------------------------------------------
    # Confidence skoru
    # ------------------------------------------------------------------

    def _calculate_confidence(self, content_json: Dict, stats: Dict) -> float:
        """
        Parse kalitesine göre 0.0–1.0 arası confidence skoru hesapla.

        Puan tablosu:
          +0.30  En az 1 ekran çıkarıldı
          +0.20  5+ ekran çıkarıldı
          +0.30  En az 1 ekranda iş kuralı var
          +0.20  İş kuralı olan ekran oranıyla ağırlıklı
        """
        score    = 0.0
        ekranlar = content_json.get("ekranlar", [])

        if not ekranlar:
            return score

        score += 0.3
        if len(ekranlar) >= 5:
            score += 0.2

        screens_with_rules = sum(
            1 for e in ekranlar
            if any(ia.get("kurallar") for ia in e.get("is_akislari", []))
        )
        if screens_with_rules > 0:
            score += 0.3

        ratio  = screens_with_rules / len(ekranlar)
        score += ratio * 0.2

        return min(score, 1.0)

    # ------------------------------------------------------------------
    # Uyarı üretimi
    # ------------------------------------------------------------------

    def _generate_warnings(self, content_json: Dict, stats: Dict) -> List[str]:
        """Parse sonucuna göre kullanıcı uyarıları oluştur."""
        warnings = []

        for ekran in content_json.get("ekranlar", []):
            has_rules = any(
                ia.get("kurallar") for ia in ekran.get("is_akislari", [])
            )
            if not has_rules:
                warnings.append(
                    f"'{ekran['ekran_adi']}' ekranında iş akışı bulunamadı"
                )

        if stats.get("links", 0) == 0:
            warnings.append("Doküman içinde hiç link bulunamadı")

        return warnings


# ---------------------------------------------------------------------------
# Unified file-type detector (DOCX + XLSX)
# ---------------------------------------------------------------------------

def detect_document_type(file_content: bytes, file_name: str) -> Dict[str, Any]:
    """
    Dosya tipini ve şablon formatını otomatik tespit eder.

    Args:
        file_content: Dosyanın binary içeriği
        file_name:    Orijinal dosya adı (uzantı tespiti için)

    Returns:
        {
            "file_type":  "xlsx" | "docx" | "unknown",
            "template":   "loodos_test_case" | "loodos_ba_bullet" |
                          "loodos_ba_table" | "generic_xlsx" | "generic",
            "confidence": 0.0–1.0,
            "details":    {...}
        }
    """
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    if ext == "xlsx":
        return _detect_xlsx_type(file_content)
    elif ext == "docx":
        return _detect_docx_type(file_content)
    else:
        return {
            "file_type":  "unknown",
            "template":   "generic",
            "confidence": 0.0,
            "details":    {"reason": f"Unsupported extension: {ext}"},
        }


def _detect_xlsx_type(file_content: bytes) -> Dict[str, Any]:
    """
    XLSX dosyasının Loodos TC formatı olup olmadığını tespit eder.

    Confidence tablosu:
      +0.2  Cover sheet mevcutsa
      +0.3  "TEST CASE ID" veya "TESTCASE" veya "TEST CASE" header'ı varsa
      +0.2  "PRIORTY" veya "PRIORITY" header'ı varsa
      +0.3  "TEST STEPS" VE "EXPECTED RESULT" header'ları varsa
      ──────
      ≥ 0.7 → loodos_test_case
      < 0.7 → generic_xlsx
    """
    try:
        import openpyxl
        # read_only=False — read_only modunda max_column/max_row None döner
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception as exc:
        return {
            "file_type":  "xlsx",
            "template":   "generic_xlsx",
            "confidence": 0.0,
            "details":    {"error": str(exc)},
        }

    sheet_names = wb.sheetnames
    score = 0.0
    details: Dict[str, Any] = {"sheet_count": len(sheet_names)}

    # Cover sheet → +0.2
    has_cover = "Cover" in sheet_names
    if has_cover:
        score += 0.2
    details["has_cover"] = has_cover

    # Revision Changes sheet → extra signal (not scored separately)
    details["has_revision"] = "Revision Changes" in sheet_names

    # Scan content sheets for known TC headers
    tc_name_found    = False
    priority_found   = False
    steps_found      = False
    expected_found   = False

    _TC_NAME_ALIASES   = {"TESTCASE", "TEST CASE", "TEST CASE ID"}
    _PRIORITY_ALIASES  = {"PRIORTY", "PRIORITY"}
    _STEPS_ALIAS       = "TEST STEPS"
    _EXPECTED_ALIAS    = "EXPECTED RESULT"

    from pipeline.tc_xlsx_reader import SKIP_SHEETS

    for sheet_name in sheet_names:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]

        # iter_rows ile ilk max_search satırı tara
        headers_upper = set()
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            non_empty = [str(v).strip().upper() for v in row if v is not None and str(v).strip()]
            if not non_empty:
                continue
            # Header satırı: bilinen alias'lardan biri içeriyor mu?
            all_aliases_upper = _TC_NAME_ALIASES | _PRIORITY_ALIASES | {_STEPS_ALIAS, _EXPECTED_ALIAS}
            if any(v in all_aliases_upper for v in non_empty) or len(non_empty) >= 3:
                headers_upper = set(non_empty)
                break

        if headers_upper & _TC_NAME_ALIASES:
            tc_name_found = True
        if headers_upper & _PRIORITY_ALIASES:
            priority_found = True
        if _STEPS_ALIAS in headers_upper:
            steps_found = True
        if _EXPECTED_ALIAS in headers_upper:
            expected_found = True

        # Found enough signals from one sheet — stop scanning
        if tc_name_found and priority_found and steps_found:
            break

    wb.close()

    if tc_name_found:
        score += 0.3
    if priority_found:
        score += 0.2
    if steps_found and expected_found:
        score += 0.3

    details.update({
        "tc_name_found":   tc_name_found,
        "priority_found":  priority_found,
        "steps_found":     steps_found,
        "expected_found":  expected_found,
        "score":           score,
    })

    template = "loodos_test_case" if score >= 0.7 else "generic_xlsx"

    return {
        "file_type":  "xlsx",
        "template":   template,
        "confidence": min(score, 1.0),
        "details":    details,
    }


def _detect_docx_type(file_content: bytes) -> Dict[str, Any]:
    """
    DOCX dosyasının şablon tipini tespit eder (mevcut DocxImportOrchestrator mantığı).
    """
    try:
        from pipeline.document_reader import read_docx_structured
        elements = read_docx_structured(file_content)
        orchestrator = DocxImportOrchestrator()
        stats    = orchestrator._calculate_stats(elements)
        template = orchestrator._detect_template(elements, stats)

        # Rough confidence estimate based on template
        confidence_map = {
            "loodos_ba_bullet": 0.85,
            "loodos_ba_table":  0.75,
            "generic":          0.40,
        }
        confidence = confidence_map.get(template, 0.4)

        return {
            "file_type":  "docx",
            "template":   template,
            "confidence": confidence,
            "details":    stats,
        }
    except Exception as exc:
        return {
            "file_type":  "docx",
            "template":   "generic",
            "confidence": 0.1,
            "details":    {"error": str(exc)},
        }
