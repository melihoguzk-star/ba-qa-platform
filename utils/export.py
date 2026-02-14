"""Export fonksiyonları: JSON, Excel (TC), Markdown."""
import json
import io
from datetime import datetime


def export_json(data: dict) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)


def export_ba_markdown(ba_data: dict, project_name: str = "") -> str:
    lines = [f"# İş Analizi — {project_name}", f"_Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}_\n"]
    moduller = ba_data.get("moduller", [])
    for i, m in enumerate(moduller, 1):
        lines.append(f"## {i}. {m.get('modul_adi', 'Modül')}")
        lines.append(m.get("aciklama", ""))
        if m.get("is_akisi"):
            lines.append("\n### İş Akışı")
            adimlar = m["is_akisi"].get("adimlar", [])
            for a in adimlar:
                lines.append(f"- {a}")
        if m.get("is_kurallari"):
            lines.append("\n### İş Kuralları")
            for k in m["is_kurallari"]:
                lines.append(f"- **{k.get('id', '')}** {k.get('kural_adi', '')}: {k.get('aciklama', '')}")
        if m.get("kabul_kriterleri"):
            lines.append("\n### Kabul Kriterleri")
            for ac in m["kabul_kriterleri"]:
                lines.append(f"- **{ac.get('id', '')}**: {ac.get('aciklama', '')}")
        lines.append("")
    return "\n".join(lines)


def export_ta_markdown(ta_data: dict, project_name: str = "") -> str:
    lines = [f"# Teknik Analiz — {project_name}", f"_Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}_\n"]
    ta = ta_data.get("teknik_analiz", ta_data)
    if ta.get("genel_tanim"):
        g = ta["genel_tanim"]
        lines.append(f"## Genel Tanım")
        lines.append(f"**Modül:** {g.get('modul_adi', '')}")
        lines.append(f"**Stack:** {', '.join(g.get('teknoloji_stack', []))}")
        lines.append(f"**Mimari:** {g.get('mimari_yaklasim', '')}\n")
    if ta.get("endpoint_detaylari"):
        lines.append("## API Endpoint Detayları")
        for ep, detail in ta["endpoint_detaylari"].items():
            lines.append(f"### `{detail.get('method', 'GET')} {ep}`")
            lines.append(f"{detail.get('aciklama', '')}\n")
    if ta.get("validasyon_kurallari"):
        lines.append("## Validasyon Kuralları")
        for v in ta["validasyon_kurallari"]:
            lines.append(f"- **{v.get('id', '')}** [{v.get('field', '')}]: {v.get('kural', '')}")
    return "\n".join(lines)


def export_tc_excel(tc_data: dict) -> bytes:
    """23 kolonlu Loodos TC şablonunda Excel oluşturur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "EXISTANCE", "CREATED BY", "DATE", "APP BUNDLE", "TEST CASE ID",
        "BR ID", "TR ID", "PRIORITY", "CHANNEL", "TESTCASE TYPE",
        "USER TYPE", "TEST AREA", "TEST SCENARIO", "TESTCASE", "TEST STEPS",
        "PRECONDITION", "TEST DATA", "EXPECTED RESULT", "POSTCONDITION",
        "ACTUAL RESULT", "STATUS", "REGRESSION CASE", "COMMENTS"
    ]

    header_fill = PatternFill(start_color="1a1d2e", end_color="1a1d2e", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin", color="555555"),
        right=Side(style="thin", color="555555"),
        top=Side(style="thin", color="555555"),
        bottom=Side(style="thin", color="555555"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    test_cases = tc_data.get("test_cases", [])
    field_map = {
        "existance": 1, "created_by": 2, "date": 3, "app_bundle": 4,
        "test_case_id": 5, "br_id": 6, "tr_id": 7, "priority": 8,
        "channel": 9, "testcase_type": 10, "user_type": 11, "test_area": 12,
        "test_scenario": 13, "testcase": 14, "test_steps": 15,
        "precondition": 16, "test_data": 17, "expected_result": 18,
        "postcondition": 19, "actual_result": 20, "status": 21,
        "regression_case": 22, "comments": 23,
    }

    for row_idx, tc in enumerate(test_cases, 2):
        for field, col in field_map.items():
            val = tc.get(field, "")
            cell = ws.cell(row=row_idx, column=col, value=str(val))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    # Kolon genişlikleri
    widths = [10, 12, 10, 15, 20, 10, 10, 10, 8, 12, 12, 15, 25, 30, 40, 20, 20, 30, 15, 15, 10, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    ws.auto_filter.ref = f"A1:W{len(test_cases) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
